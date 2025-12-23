"""
Excel output generator for VRP routing solutions.
Generates professional Excel files with route details and summary.
"""
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.route import RoutingSolution, Route, RouteStop
from ..models.location import Depot


class ExcelGenerator:
    """
    Generates Excel output files for VRP routing solutions.

    Creates a professional Excel workbook with two sheets:
    1. "Routes by Vehicle" - Detailed route information grouped by vehicle
    2. "Summary" - High-level metrics and solution overview
    """

    # Color definitions
    COLOR_HEADER = "366092"  # Dark blue
    COLOR_PRIORITY_HIGH = "FFD966"  # Yellow
    COLOR_PRIORITY_CRITICAL = "FF6B6B"  # Red
    COLOR_SUBTOTAL = "E2EFDA"  # Light green
    COLOR_SUMMARY_HEADER = "4472C4"  # Blue

    def __init__(self, depot: Depot):
        """
        Initialize the Excel generator.

        Args:
            depot: Depot location for route information
        """
        self.depot = depot
        self.wb: Optional[Workbook] = None

    def generate(
        self,
        solution: RoutingSolution,
        output_dir: str = "results",
        filename: Optional[str] = None
    ) -> Path:
        """
        Generate Excel file for the routing solution.

        Args:
            solution: RoutingSolution object containing all routes
            output_dir: Directory to save the Excel file (default: "results")
            filename: Optional custom filename (without extension)

        Returns:
            Path to the generated Excel file

        Raises:
            ValueError: If solution has no routes
        """
        if not solution.routes or solution.total_vehicles_used == 0:
            raise ValueError("Cannot generate Excel for empty solution")

        # Create output directory if it doesn't exist
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # Generate filename with timestamp
        if filename is None:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"routing_result_{timestamp}"

        filepath = output_path / f"{filename}.xlsx"

        # Create workbook
        self.wb = Workbook()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Generate sheets
        self._generate_routes_sheet(solution)
        self._generate_summary_sheet(solution)

        # Save workbook
        self.wb.save(filepath)

        return filepath

    def _generate_routes_sheet(self, solution: RoutingSolution):
        """
        Generate the "Routes by Vehicle" sheet.

        Args:
            solution: RoutingSolution object
        """
        ws = self.wb.create_sheet("Routes by Vehicle")

        # Define column headers
        headers = [
            "Source",                   # 1
            "Trip #",                   # 2
            "Vehicle Name",             # 3
            "Delivery Time",            # 4
            "Customer",                 # 5
            "Address",                  # 6
            "Kelurahan",                # 7
            "Kecamatan",                # 8
            "City",                     # 9
            "Rate (Rp/km)",             # 10
            "Weight (kg)",              # 11
            "Arrival Time",             # 12
            "Departure Time",           # 13
            "Distance (km)",            # 14
            "Cost (Rp)",                # 15 (NEW)
            "Cumulative Weight (kg)",   # 16
            "Sequence",                 # 17
            "Latitude",                 # 18
            "Longitude",                # 19
            "Notes"                     # 20
        ]

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._get_border()

        # Set column widths (20 columns now)
        column_widths = [10, 8, 18, 14, 25, 35, 15, 15, 15, 12, 12, 13, 13, 12, 15, 18, 10, 12, 12, 20]
        for col_num, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Write route data
        current_row = 2

        # Filter out empty routes
        active_routes = [r for r in solution.routes if r.num_stops > 0]

        for route in active_routes:
            start_row = current_row

            # Write each stop in the route
            for stop in route.stops:
                self._write_route_stop_row(ws, current_row, route, stop)
                current_row += 1

            # Write subtotal row
            self._write_subtotal_row(ws, current_row, route, start_row, current_row - 1)
            current_row += 1

            # Add grouping/outline for this vehicle
            if current_row - start_row > 1:  # Only group if there are stops
                ws.row_dimensions.group(start_row, current_row - 2, hidden=False)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _write_route_stop_row(self, ws: Worksheet, row: int, route: Route, stop: RouteStop):
        """
        Write a single route stop row.

        Args:
            ws: Worksheet object
            row: Row number to write
            route: Route object
            stop: RouteStop object
        """
        order = stop.order

        # Column values (20 columns)
        values = [
            route.source,  # A: Source (HUB or DEPOT)
            route.trip_number,  # B: Trip #
            route.vehicle.name,  # C: Vehicle Name
            order.delivery_time,  # D: Delivery Time
            order.display_name,  # E: Customer
            order.alamat,  # F: Address
            order.kelurahan or "",  # G: Kelurahan
            order.kecamatan or "",  # H: Kecamatan
            order.kota or "",  # I: City
            route.vehicle.cost_per_km,  # J: Rate (Rp/km)
            order.load_weight_in_kg,  # K: Weight (kg)
            stop.arrival_time_str,  # L: Arrival Time
            stop.departure_time_str,  # M: Departure Time
            stop.distance_from_prev,  # N: Distance (km)
            stop.distance_from_prev * route.vehicle.cost_per_km,  # O: Cost (Rp) - NEW
            stop.cumulative_weight,  # P: Cumulative Weight (kg)
            stop.sequence,  # Q: Sequence
            order.coordinates[0],  # R: Latitude
            order.coordinates[1],  # S: Longitude
            "PRIORITY" if order.is_priority else ""  # T: Notes
        ]

        # Write values
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = self._get_border(thin=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Format specific columns
            if col_num == 10:  # Rate - currency
                cell.number_format = 'Rp #,##0'
            elif col_num == 15:  # Cost - currency (NEW)
                cell.number_format = 'Rp #,##0'
            elif col_num in [11, 14, 16]:  # Weight, Distance, Cumulative Weight - numbers
                cell.number_format = '0.00'
            elif col_num == 17:  # Sequence - integer
                cell.number_format = '0'
            elif col_num in [18, 19]:  # Coordinates - decimal
                cell.number_format = '0.000000'

        # Apply priority color coding
        if order.is_priority:
            fill_color = self.COLOR_PRIORITY_HIGH
            for col_num in range(1, 21):  # 20 columns
                ws.cell(row=row, column=col_num).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )

    def _write_subtotal_row(self, ws: Worksheet, row: int, route: Route,
                           start_row: int, end_row: int):
        """
        Write a subtotal row for a vehicle route.

        Args:
            ws: Worksheet object
            row: Row number to write
            route: Route object
            start_row: First data row for this route
            end_row: Last data row for this route
        """
        # Calculate totals
        total_cost = route.total_cost
        total_weight = route.total_weight
        total_distance = route.total_distance

        # Write subtotal cells
        ws.cell(row=row, column=1).value = f"SUBTOTAL - {route.vehicle.name}"
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)

        ws.cell(row=row, column=3).value = f"{route.num_stops} stops"
        ws.cell(row=row, column=11).value = total_weight
        ws.cell(row=row, column=11).number_format = '0.00'
        ws.cell(row=row, column=14).value = total_distance
        ws.cell(row=row, column=14).number_format = '0.00'

        # Cost total (now in column 15 - aligned with "Cost (Rp)" header)
        ws.cell(row=row, column=15).value = total_cost
        ws.cell(row=row, column=15).number_format = 'Rp #,##0'

        # Apply subtotal formatting
        for col_num in range(1, 21):  # 20 columns
            cell = ws.cell(row=row, column=col_num)
            cell.fill = PatternFill(start_color=self.COLOR_SUBTOTAL,
                                   end_color=self.COLOR_SUBTOTAL,
                                   fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = self._get_border()

    def _generate_summary_sheet(self, solution: RoutingSolution):
        """
        Generate the "Summary" sheet.

        Args:
            solution: RoutingSolution object
        """
        ws = self.wb.create_sheet("Summary")

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

        # Title
        ws['A1'] = "ROUTING SOLUTION SUMMARY"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_SUMMARY_HEADER,
                                   end_color=self.COLOR_SUMMARY_HEADER,
                                   fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:B1')

        # Summary data
        summary_data = [
            ("", ""),  # Empty row
            ("Generation Details", ""),
            ("Generated Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Optimization Strategy", solution.optimization_strategy.replace("_", " ").title()),
            ("Computation Time", f"{solution.computation_time:.2f} seconds"),
            ("", ""),  # Empty row
            ("Depot Information", ""),
            ("Depot Name", self.depot.name),
            ("Depot Address", self.depot.address if hasattr(self.depot, 'address') else "N/A"),
            ("Depot Coordinates", f"{self.depot.coordinates[0]:.6f}, {self.depot.coordinates[1]:.6f}"),
            ("", ""),  # Empty row
            ("Route Metrics", ""),
            ("Total Vehicles Used", solution.total_vehicles_used),
            ("Total Orders Delivered", solution.total_orders_delivered),
            ("Unassigned Orders", len(solution.unassigned_orders)),
            ("Total Distance", f"{solution.total_distance:.2f} km"),
            ("Total Cost", f"Rp {solution.total_cost:,.0f}"),
            ("Average Cost per Vehicle", f"Rp {solution.total_cost / max(solution.total_vehicles_used, 1):,.0f}"),
            ("Average Distance per Vehicle", f"{solution.total_distance / max(solution.total_vehicles_used, 1):.2f} km"),
        ]

        # Write summary data
        current_row = 2
        for label, value in summary_data:
            if label == "":  # Empty row
                current_row += 1
                continue

            cell_a = ws.cell(row=current_row, column=1)
            cell_b = ws.cell(row=current_row, column=2)

            cell_a.value = label
            cell_b.value = value

            # Check if this is a section header
            if value == "":
                cell_a.font = Font(bold=True, size=12, color="FFFFFF")
                cell_a.fill = PatternFill(start_color=self.COLOR_HEADER,
                                         end_color=self.COLOR_HEADER,
                                         fill_type="solid")
                ws.merge_cells(f'A{current_row}:B{current_row}')
            else:
                cell_a.font = Font(bold=True, size=11)
                cell_b.font = Font(size=11)
                cell_a.alignment = Alignment(horizontal="left", vertical="center")
                cell_b.alignment = Alignment(horizontal="left", vertical="center")

            cell_a.border = self._get_border(thin=True)
            cell_b.border = self._get_border(thin=True)

            current_row += 1

        # Add vehicle breakdown section
        ws.cell(row=current_row, column=1).value = "Vehicle Breakdown"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=self.COLOR_HEADER,
                                                              end_color=self.COLOR_HEADER,
                                                              fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.cell(row=current_row, column=1).border = self._get_border(thin=True)
        ws.cell(row=current_row, column=2).border = self._get_border(thin=True)
        current_row += 1

        # Vehicle details headers (now includes utilization)
        vehicle_headers = ["Vehicle Name", "Stops", "Weight (kg)", "Capacity (kg)", "Utilization %", "Distance (km)", "Cost (Rp)"]
        for col_num, header in enumerate(vehicle_headers, start=1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()

        # Adjust column widths for vehicle breakdown
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18

        current_row += 1

        # Write vehicle details with utilization metrics
        active_routes = [r for r in solution.routes if r.num_stops > 0]
        total_weight_used = 0.0
        total_capacity_available = 0.0

        for route in active_routes:
            weight = route.total_weight
            capacity = route.vehicle.capacity
            utilization = (weight / capacity * 100) if capacity > 0 else 0

            total_weight_used += weight
            total_capacity_available += capacity

            ws.cell(row=current_row, column=1).value = route.vehicle.name
            ws.cell(row=current_row, column=2).value = route.num_stops
            ws.cell(row=current_row, column=3).value = weight
            ws.cell(row=current_row, column=3).number_format = '0.00'
            ws.cell(row=current_row, column=4).value = capacity
            ws.cell(row=current_row, column=4).number_format = '0.00'
            ws.cell(row=current_row, column=5).value = utilization
            ws.cell(row=current_row, column=5).number_format = '0.0"%"'
            ws.cell(row=current_row, column=6).value = route.total_distance
            ws.cell(row=current_row, column=6).number_format = '0.00'
            ws.cell(row=current_row, column=7).value = route.total_cost
            ws.cell(row=current_row, column=7).number_format = 'Rp #,##0'

            # Color-code utilization (green=good, yellow=medium, red=low)
            util_cell = ws.cell(row=current_row, column=5)
            if utilization >= 70:
                util_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            elif utilization >= 40:
                util_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
            else:
                util_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

            for col_num in range(1, 8):
                ws.cell(row=current_row, column=col_num).border = self._get_border(thin=True)
                ws.cell(row=current_row, column=col_num).alignment = Alignment(vertical="center")

            current_row += 1

        # Add fleet utilization summary row
        if active_routes:
            overall_utilization = (total_weight_used / total_capacity_available * 100) if total_capacity_available > 0 else 0

            ws.cell(row=current_row, column=1).value = "FLEET TOTAL"
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=3).value = total_weight_used
            ws.cell(row=current_row, column=3).number_format = '0.00'
            ws.cell(row=current_row, column=4).value = total_capacity_available
            ws.cell(row=current_row, column=4).number_format = '0.00'
            ws.cell(row=current_row, column=5).value = overall_utilization
            ws.cell(row=current_row, column=5).number_format = '0.0"%"'
            ws.cell(row=current_row, column=6).value = solution.total_distance
            ws.cell(row=current_row, column=6).number_format = '0.00'
            ws.cell(row=current_row, column=7).value = solution.total_cost
            ws.cell(row=current_row, column=7).number_format = 'Rp #,##0'

            for col_num in range(1, 8):
                cell = ws.cell(row=current_row, column=col_num)
                cell.fill = PatternFill(start_color=self.COLOR_SUBTOTAL, end_color=self.COLOR_SUBTOTAL, fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = self._get_border()

            current_row += 1

        # Add unassigned orders section if any
        if solution.unassigned_orders:
            current_row += 1
            ws.cell(row=current_row, column=1).value = "⚠️ UNASSIGNED ORDERS"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FF0000",
                                                                  end_color="FF0000",
                                                                  fill_type="solid")
            ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1

            for order in solution.unassigned_orders:
                ws.cell(row=current_row, column=1).value = order.sale_order_id
                ws.cell(row=current_row, column=2).value = order.display_name
                current_row += 1

    def _get_border(self, thin: bool = False) -> Border:
        """
        Get a border style for cells.

        Args:
            thin: If True, use thin borders; otherwise medium

        Returns:
            Border object
        """
        style = "thin" if thin else "medium"
        return Border(
            left=Side(style=style),
            right=Side(style=style),
            top=Side(style=style),
            bottom=Side(style=style)
        )
